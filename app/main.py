"""
产品目录自动生成工具 - 内部使用
FastAPI 后端：Excel 上传 → 数据预览 → PDF 目录 → PNG 总览图

部署: Railway (Docker)
"""
from fastapi import FastAPI, UploadFile, File, Request, Query
from fastapi.responses import HTMLResponse, Response
from fastapi.templating import Jinja2Templates
import openpyxl
import io
import os
import base64
import tempfile
from datetime import datetime
from pathlib import Path

# QR 码生成
import qrcode as qrcode_lib

# Playwright 用于 PDF / PNG 渲染
from playwright.async_api import async_playwright

app = FastAPI(title="产品目录自动生成工具")

# 模板目录
BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# 上传 & 输出目录
UPLOADS_DIR = BASE_DIR / "static" / "uploads"
OUTPUTS_DIR = BASE_DIR / "static" / "outputs"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════
#  Excel 字段定义
# ═══════════════════════════════════════════════════════════════

TARGET_FIELDS = [
    "Product Name",
    "Brand",
    "Price",
    "Category",
    "Weight",
    "Dimensions",
    "Selling Point 1",
    "Selling Point 2",
    "Selling Point 3",
    "Promotion Angle",
    "Product URL",
    "Image URL",
]

# 中文 → 英文 字段名映射（同时支持中英文表头）
FIELD_ALIAS_MAP: dict[str, str] = {
    # 中文表头别名
    "商品标题": "Product Name",
    "品牌": "Brand",
    "价格($)": "Price",
    "价格": "Price",
    "price": "Price",
    "大类目": "Category",
    "类目": "Category",
    "商品重量（单位换算）": "Weight",
    "商品重量": "Weight",
    "商品尺寸（单位换算）": "Dimensions",
    "商品尺寸": "Dimensions",
    "产品卖点": "Selling Points",  # 特殊处理：合并单元格
    "商品详情页链接": "Product URL",
    "商品主图": "Image URL",
    "图片链接": "Image URL",
    "图片": "Image URL",
    "推广方向": "Promotion Angle",
    "推广角度": "Promotion Angle",
    # 覆盖列（如果英文名和中文表头都不匹配）
    "详细参数": "_specs",
}

# 需要忽略的列（序号之类的）
IGNORE_HEADERS = {"#", "序号", "No.", "No", "编号"}

# ═══════════════════════════════════════════════════════════════
#  服务端数据存储
# ═══════════════════════════════════════════════════════════════

_server_state: dict = {
    "products": None,
    "filename": None,
    "excel_path": None,
}


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ═══════════════════════════════════════════════════════════════
#  类目分组
# ═══════════════════════════════════════════════════════════════

def group_by_category(products: list[dict]) -> list[tuple[str, list[dict]]]:
    """按 Category 字段分组，保持原始出现顺序。空类目归入 'Other'。"""
    groups: dict[str, list[dict]] = {}
    order: list[str] = []
    for p in products:
        cat = p.get("Category", "").strip() or "Other"
        if cat not in groups:
            groups[cat] = []
            order.append(cat)
        groups[cat].append(p)
    return [(cat, groups[cat]) for cat in order]


# ═══════════════════════════════════════════════════════════════
#  Excel 解析
# ═══════════════════════════════════════════════════════════════

def parse_excel(file_bytes: bytes) -> list[dict]:
    """
    解析 Excel 文件，逐行读取产品数据。
    
    支持中英文表头自动识别 (FIELD_ALIAS_MAP)。
    产品卖点 (Selling Points) 支持合并单元格（\\n 分割 → Selling Point 1/2/3）。
    缺失字段自动留空，跳过全空行和序号列。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 读取表头
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    # 建立字段名 → 列索引的映射 (先匹配英文原名, 再匹配中文别名)
    field_col_map: dict[str, int] = {}
    has_combined_selling_points = False
    selling_points_col = -1
    specs_col = -1

    for idx, header in enumerate(headers):
        if not header:
            continue
        # 跳过忽略列
        if header in IGNORE_HEADERS:
            continue
        # 精确匹配 TARGET_FIELDS
        if header in TARGET_FIELDS:
            field_col_map[header] = idx
            continue
        # 中文别名匹配
        if header in FIELD_ALIAS_MAP:
            mapped = FIELD_ALIAS_MAP[header]
            if mapped == "Selling Points":
                has_combined_selling_points = True
                selling_points_col = idx
            elif mapped == "_specs":
                specs_col = idx
            else:
                field_col_map[mapped] = idx
            continue
        # 模糊匹配：header 包含已知字段名
        for en_field in TARGET_FIELDS:
            if en_field.lower() in header.lower():
                field_col_map[en_field] = idx
                break

    # 逐行读取
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过全空行
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        product = {}
        for field in TARGET_FIELDS:
            if field in field_col_map:
                cell_value = row[field_col_map[field]]
                product[field] = str(cell_value).strip() if cell_value is not None else ""
            else:
                product[field] = ""

        # 处理合并卖点列 (\\n 分割 → Selling Point 1/2/3)
        if has_combined_selling_points and selling_points_col >= 0:
            raw = row[selling_points_col]
            if raw and str(raw).strip():
                lines = [l.strip() for l in str(raw).split("\n") if l.strip()]
                for i in range(3):
                    key = f"Selling Point {i + 1}"
                    product[key] = lines[i] if i < len(lines) else ""

        # 从详细参数中提取推广方向 (取前两行作为 Promo)
        if specs_col >= 0 and not product.get("Promotion Angle"):
            raw = row[specs_col]
            if raw and str(raw).strip():
                first_line = str(raw).split("\n")[0].strip()
                # 去掉 Brand: 等前缀
                if ":" in first_line:
                    first_line = first_line.split(":", 1)[1].strip()
                product["Promotion Angle"] = first_line[:120]

        # 跳过全空产品
        if any(v.strip() for v in product.values() if v):
            products.append(product)

    wb.close()
    return products


# ═══════════════════════════════════════════════════════════════
#  QR 码生成
# ═══════════════════════════════════════════════════════════════

def generate_qr_codes(products: list[dict]) -> dict[int, str]:
    """为有 Product URL 的产品生成 QR 码 base64 图片"""
    qr_codes = {}
    for idx, p in enumerate(products):
        url = p.get("Product URL", "")
        if url and url.strip():
            try:
                qr = qrcode_lib.QRCode(
                    version=1,
                    error_correction=qrcode_lib.constants.ERROR_CORRECT_M,
                    box_size=10,
                    border=1,
                )
                qr.add_data(url.strip())
                qr.make(fit=True)
                img = qr.make_image(fill_color="#1a2744", back_color="#ffffff")
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                qr_codes[idx] = base64.b64encode(buf.getvalue()).decode("utf-8")
            except Exception:
                pass
    return qr_codes


# ═══════════════════════════════════════════════════════════════
#  PDF 生成 (Playwright Chromium)
# ═══════════════════════════════════════════════════════════════

async def render_pdf(products: list[dict], qr_codes: dict[int, str]) -> bytes:
    """Jinja2 渲染 catalog.html → Playwright Chromium 导出 PDF"""
    categories = list(dict.fromkeys(
        p.get("Category", "") for p in products if p.get("Category", "")
    ))[:6]
    grouped = group_by_category(products)

    html_content = templates.get_template("catalog.html").render(
        request=None,
        products=products,
        grouped_products=grouped,
        qr_codes=qr_codes,
        categories=categories,
        now=datetime.now(),
    )

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            args=["--no-sandbox", "--disable-setuid-sandbox"]
        )
        page = await browser.new_page()
        await page.set_content(html_content, wait_until="networkidle", timeout=30000)
        await page.wait_for_timeout(2000)

        pdf_bytes = await page.pdf(
            format="A4",
            landscape=True,
            print_background=True,
            margin={"top": "0", "bottom": "0", "left": "0", "right": "0"},
            prefer_css_page_size=True,
        )
        await browser.close()
        return pdf_bytes


# ═══════════════════════════════════════════════════════════════
#  PNG 生成 (Playwright Chromium 截图)
# ═══════════════════════════════════════════════════════════════

async def render_overview_png(products: list[dict]) -> bytes:
    """Jinja2 渲染 overview.html → Playwright Chromium 全页截图 PNG"""
    grouped = group_by_category(products)
    html_content = templates.get_template("overview.html").render(
        request=None,
        products=products,
        grouped_products=grouped,
    )

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            args=["--no-sandbox", "--disable-setuid-sandbox"]
        )
        page = await browser.new_page(viewport={"width": 800, "height": 600})
        await page.set_content(html_content, wait_until="networkidle", timeout=30000)
        await page.wait_for_timeout(3000)

        body_height = await page.evaluate("() => document.body.scrollHeight")
        await page.set_viewport_size({"width": 800, "height": body_height + 20})

        png_bytes = await page.screenshot(full_page=True, type="png")
        await browser.close()
        return png_bytes


# ═══════════════════════════════════════════════════════════════
#  路由
# ═══════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """主页"""
    return templates.TemplateResponse("index.html", {
        "request": request,
        "products": _server_state.get("products"),
        "error": None,
        "filename": _server_state.get("filename"),
    })


@app.post("/upload", response_class=HTMLResponse)
async def upload_excel(request: Request, file: UploadFile = File(...)):
    """上传 Excel → 解析 → 存入服务端状态 → 返回预览"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return templates.TemplateResponse("index.html", {
            "request": request,
            "products": None,
            "error": "请上传 .xlsx 格式的 Excel 文件",
            "filename": None,
        })

    try:
        contents = await file.read()

        # 保存上传文件（时间戳防覆盖）
        ts = _timestamp()
        safe_name = f"upload_{ts}.xlsx"
        upload_path = UPLOADS_DIR / safe_name
        upload_path.write_bytes(contents)
        _server_state["excel_path"] = str(upload_path)

        products = parse_excel(contents)

        if not products:
            return templates.TemplateResponse("index.html", {
                "request": request,
                "products": None,
                "error": "Excel 文件中未识别到任何产品数据，请检查表头字段名是否正确",
                "filename": file.filename,
            })

        _server_state["products"] = products
        _server_state["filename"] = file.filename

        return templates.TemplateResponse("index.html", {
            "request": request,
            "products": products,
            "error": None,
            "filename": file.filename,
        })

    except Exception as e:
        return templates.TemplateResponse("index.html", {
            "request": request,
            "products": None,
            "error": f"解析 Excel 文件时出错：{str(e)}",
            "filename": file.filename,
        })


@app.get("/generate-pdf")
async def generate_pdf(inline: str = Query(default="0")):
    """生成 PDF 产品目录。?inline=1 浏览器预览，否则下载。"""
    products = _server_state.get("products")
    if not products:
        return Response(content="No data. Please upload an Excel file first.", status_code=400)

    qr_codes = generate_qr_codes(products)

    try:
        pdf_bytes = await render_pdf(products, qr_codes)
    except Exception as e:
        return Response(content=f"PDF generation failed: {str(e)}", status_code=500)

    # 保存输出文件
    output_path = OUTPUTS_DIR / f"product_catalog_{_timestamp()}.pdf"
    output_path.write_bytes(pdf_bytes)

    disposition = "inline" if inline == "1" else "attachment"
    filename = f"product_catalog_{_timestamp()}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
        },
    )


@app.get("/generate-overview")
async def generate_overview_png():
    """生成产品总览宣传图 PNG (800px 宽竖版海报)"""
    products = _server_state.get("products")
    if not products:
        return Response(content="No data. Please upload an Excel file first.", status_code=400)

    try:
        png_bytes = await render_overview_png(products)
    except Exception as e:
        return Response(content=f"PNG generation failed: {str(e)}", status_code=500)

    # 保存输出文件
    output_path = OUTPUTS_DIR / f"product_overview_{_timestamp()}.png"
    output_path.write_bytes(png_bytes)

    filename = f"product_overview_{_timestamp()}.png"
    return Response(
        content=png_bytes,
        media_type="image/png",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(png_bytes)),
        },
    )


# ═══════════════════════════════════════════════════════════════
#  入口
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("app.main:app", host="0.0.0.0", port=port, reload=True)
